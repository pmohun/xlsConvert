import xlrd
import os
from tkinter import filedialog 
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException

# Function to get ONLY the .xls files from cwd
# and convert them to .xlsx files via open xls2xlsx(src_file)

dirName = filedialog.askdirectory() 
os.chdir(dirName)

def main():
    for filename in os.listdir(dirName): #os.getcwd()):
        if filename.endswith('.xls'):
            xls2xlsx(filename)

# Function to convert a SINGLE .xls workbook 
# to a .xlsx workbook given a src_file_path
def xls2xlsx(filename):
    # Old .xls workbook
    book_xls = xlrd.open_workbook(filename)
    # New .xlsx workbook
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    # Iterate through .xls sheet names to create matching ones in new workbook
    for sheet_index in range(0,len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
	# If its the first sheet then get the first .xlsx sheet
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.get_active_sheet()
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

	# Nested for loop to go through each cell and copy over to the new .xlsx sheet
        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row + 1 , column = col + 1).value = sheet_xls.cell_value(row, col)

    # After going through each sheet and copying them over to .xlsx 
    # save the new .xlsx workbook by appending 'x' to file extension
    book_xlsx.save(filename + "x")



if __name__ == "__main__":
    main()