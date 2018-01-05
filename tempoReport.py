#! python3	
# tempoReport.py - Pulls timesheet information for consolidation into report

import openpyxl, time, os, shutil
import tkinter as tk
from tkinter import filedialog
from collections import defaultdict
# from xls2xlsx import xls2xlsx
root = tk.Tk()
root.withdraw()

##print('Select folder containing files to convert. . .')
##dirName = filedialog.askdirectory() 
# xls2xlsx(dirName) # convert files to xlsx

# choose workbook to write to
dirName = filedialog.askdirectory() 
file_path = dirName
output = openpyxl.load_workbook(file_path)
timeData = output.get_sheet_by_name('Time Data')
approvalData = output.get_sheet_by_name('Approval Data')

# Parse data extract folder, pull data and consolidate
print('Select folder containing .xlsx files to read. . .')
dirName = filedialog.askdirectory()
os.chdir(dirName)
for root, dirs, file in os.walk(dirName):
    #print(file)
    for filename in file:
#file_path = filedialog.askopenfilename() # delete in final version
        filename = str(filename)
        wb = openpyxl.load_workbook(filename)
        worklogs = wb.get_sheet_by_name('Worklogs')
        people = wb.get_sheet_by_name('People')
        max_row_worklogs = worklogs.max_row
        max_row_people = people.max_row
        #max_row_approvalData = approvalData.max_row+1
        #max_row_timeData = timeData.max_row + 1
        people_range = 'H' + str(max_row_people)
        worklogs_range = 'F' + str(max_row_worklogs)
        #Pull approval data and write to new report
        for row in people.iter_rows('A1:H'+str(max_row_people)):
            week = people['H1'].value
            max_row_approvalData = approvalData.max_row + 1
            for cell in row:
                if cell.column == 'A':   
                    coordinates = cell.column + str(max_row_approvalData)
                    approvalData[coordinates] = cell.value
                elif cell.column == 'B':
                    coordinates = cell.column + str(max_row_approvalData)
                    approvalData[coordinates] = cell.value
                elif cell.column == 'G':
                    coordinates = 'C' + str(max_row_approvalData)
                    approvalData[coordinates] = cell.value
                    approvalData['D'+ str(max_row_approvalData)] = week
        # Pull worklog data and write to new report
        for row in worklogs.iter_rows('A2:'+worklogs_range):
            max_row_timeData = timeData.max_row + 1
            print('processing row %d' % max_row_timeData)
            for cell in row:
                coordinates = cell.column + str(max_row_timeData)
                timeData[coordinates] = cell.value

print('Where should I save this report?')
dirName = filedialog.askdirectory()
os.chdir(dirName)
output.save('Tempo_Report_'+str(time.strftime("%m_%d_%Y"))+'.xlsx')